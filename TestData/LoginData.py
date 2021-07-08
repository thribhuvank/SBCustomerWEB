import openpyxl


class LoginData:
    test_login_data = [{"username": "8971393466", "password": "123456"}, {"username": "8971393466", "password": "123456"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:/Users/User/PycharmProjects/SBCustomerFunctional/DataFiles/loginData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
